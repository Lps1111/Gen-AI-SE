# export_excel.py
from pathlib import Path
from openpyxl import Workbook
from openpyxl.utils import get_column_letter


def export_to_excel(rows: list[dict], out_path: Path) -> None:
    out_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "Headlines"

    headers = ["Source", "Title", "URL", "Collected At"]
    ws.append(headers)

    for r in rows:
        ws.append([r["source"], r["title"], r["url"], r["collected_at"]])

    # Auto-size columns (simple)
    for col_idx in range(1, len(headers) + 1):
        col_letter = get_column_letter(col_idx)
        max_len = 12
        for cell in ws[col_letter]:
            if cell.value:
                max_len = max(max_len, min(len(str(cell.value)), 80))
        ws.column_dimensions[col_letter].width = max_len + 2

    wb.save(out_path)
